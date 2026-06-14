from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Header
from sqlmodel import Session, select
from typing import List, Optional
from pydantic import BaseModel
from database import get_session
from models import Supplier, SupplierMessage
from datetime import datetime
import json
import os
import uuid


router = APIRouter(prefix="/api/suppliers", tags=["suppliers"])


# ── Request models for Add Supplier ──────────────────────────────────
class ProductRow(BaseModel):
    product_type: str = "skincare"
    sku: str = "SKU-001"
    price: float = 50.0
    availability: float = 80.0
    number_sold: int = 500
    revenue: float = 25000.0
    customer_demographics: str = "Female"
    stock_level: float = 60.0
    lead_time: float = 15.0
    order_quantity: int = 500
    shipping_time: float = 5.0
    shipping_cost: float = 6.0
    shipping_carrier: str = "Carrier A"
    production_volume: int = 800
    manufacturing_lead_time: float = 12.0
    manufacturing_cost: float = 30.0
    defect_rate: float = 2.0
    transportation_mode: str = "Road"
    route: str = "Route A"
    inspection_result: str = "Pass"  # Pass | Fail | Pending


class AddSupplierRequest(BaseModel):
    name: str
    location: str
    products: List[ProductRow]

@router.get("/", response_model=dict)
def get_suppliers(
    session: Session = Depends(get_session),
    page: int = 1,
    limit: int = 10,
    search: Optional[str] = None,
    sortBy: Optional[str] = None,
    sortOrder: Optional[str] = "asc",
    riskLevel: Optional[str] = None,
    location: Optional[str] = None,
):
    query = select(Supplier)

    if search:
        query = query.where(
            (Supplier.name.contains(search)) |
            (Supplier.location.contains(search)) |
            (Supplier.supplier_id.contains(search))
        )
    
    if riskLevel:
        query = query.where(Supplier.risk_level == riskLevel)
    
    if location:
        query = query.where(Supplier.location.contains(location))

    # Sorting
    if sortBy:
        field = getattr(Supplier, sortBy, None)
        if field:
            if sortOrder == "desc":
                query = query.order_by(field.desc())
            else:
                query = query.order_by(field.asc())

    # Pagination
    offset = (page - 1) * limit
    suppliers = session.exec(query.offset(offset).limit(limit)).all()
    
    # Total count
    total_items = len(session.exec(query).all())
    total_pages = (total_items + limit - 1) // limit

    return {
        "success": True,
        "data": {
            "suppliers": suppliers,
            "pagination": {
                "currentPage": page,
                "totalPages": total_pages,
                "totalItems": total_items,
                "itemsPerPage": limit
            }
        }
    }

@router.get("/{supplier_id}", response_model=dict)
def get_supplier(supplier_id: str, session: Session = Depends(get_session)):
    supplier = session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"success": True, "data": supplier}

@router.get("/{supplier_id}/summary", response_model=dict)
def get_supplier_summary(supplier_id: str, session: Session = Depends(get_session)):
    """Generate an AI performance summary for a supplier."""
    supplier = session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    prompt = f"""You are a supply chain performance analyst. Generate a concise performance summary for this supplier:
    Name: {supplier.name}
    Location: {supplier.location}
    Product Types: {supplier.product_types}
    Overall Score: {supplier.overall_score or 'Not evaluated'}
    Risk Level: {supplier.risk_level or 'Not evaluated'}
    On-Time Delivery: {supplier.otd_percentage or 'Not evaluated'}%
    Defect Rate: {supplier.defect_rate}%
    Inspection Pass Rate: {supplier.inspection_pass_rate}%
    Avg Lead Time: {supplier.avg_lead_time} days
    Avg Shipping Time: {supplier.avg_shipping_time} days
    Avg Shipping Cost: ${supplier.avg_shipping_cost}
    Avg Manufacturing Cost: ${supplier.avg_manufacturing_cost}
    Avg Manufacturing Lead Time: {supplier.avg_manufacturing_lead_time} days
    Total Revenue: ${supplier.total_revenue}
    Total Products Sold: {supplier.total_products_sold}
    Production Volume: {supplier.total_production_volume}
    Avg Availability: {supplier.avg_availability}
    Transportation Modes: {supplier.transportation_modes}
    Shipping Carriers: {supplier.shipping_carriers}
    Routes: {supplier.routes}
    Avg Total Cost: ${supplier.avg_total_cost}
    Number of SKUs: {supplier.num_skus}

    Return a JSON object with:
    - summary_text: 2-3 sentence executive summary
    - key_insights: array of 4 specific data-driven insights
    - risk_flags: array of 2-3 risk flags or concerns
    - data_sources_used: array of data source names used

    Output ONLY valid JSON, no markdown formatting."""

    try:
        from llm import get_llm
        from langchain_core.messages import HumanMessage
        import os
        import json

        llm = get_llm(temperature=0.3)

        message = HumanMessage(content=prompt)
        response = llm.invoke([message])
        content = response.content

        if "```json" in content:
            content = content.split("```json")[1].split("```")[0]
        elif "```" in content:
            content = content.split("```")[1].split("```")[0]

        data = json.loads(content.strip())

        return {
            "success": True,
            "data": {
                "supplier_id": supplier.supplier_id,
                "summary_text": data.get("summary_text", ""),
                "key_insights": data.get("key_insights", []),
                "risk_flags": data.get("risk_flags", []),
                "data_sources_used": data.get("data_sources_used", ["Database"]),
                "generated_date": datetime.utcnow().isoformat()
            }
        }
    except Exception as e:
        print(f"Error generating summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{supplier_id}/report", response_model=dict)
def generate_report(supplier_id: str, session: Session = Depends(get_session)):
    supplier = session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    prompt = f"""
    Generate a concise risk assessment report for the following supplier:
    Name: {supplier.name}
    Location: {supplier.location}
    Product Types: {supplier.product_types}
    Overall Score: {supplier.overall_score}
    Risk Level: {supplier.risk_level}
    OTD Percentage: {supplier.otd_percentage}%
    Defect Rate: {supplier.defect_rate}%
    Inspection Pass Rate: {supplier.inspection_pass_rate}%
    Avg Lead Time: {supplier.avg_lead_time} days
    Avg Shipping Time: {supplier.avg_shipping_time} days
    Avg Manufacturing Cost: ${supplier.avg_manufacturing_cost}
    Total Revenue: ${supplier.total_revenue}
    
    The report should include:
    1. Executive Summary
    2. Key Risks
    3. Recommendations
    
    Format the output as a JSON object with keys: summary_text, key_insights (list), risk_flags (list), data_sources_used (list).
    Do NOT return markdown code blocks, just the JSON string.
    """

    try:
        from llm import get_llm
        from langchain_core.messages import HumanMessage
        import os
        import json
        
        llm = get_llm()
        
        message = HumanMessage(content=prompt)
        response = llm.invoke([message])
        
        content = response.content
        
        try:
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            elif "```" in content:
                content = content.split("```")[1].split("```")[0]
                
            data = json.loads(content.strip())
            
            report_content = data.get("summary_text", "Summary not generated.")
            key_insights = data.get("key_insights", [])
            risk_flags = data.get("risk_flags", [])
            data_sources_used = data.get("data_sources_used", ["Database"])
            
        except json.JSONDecodeError:
            report_content = content
            key_insights = ["Could not parse structured insights."]
            risk_flags = ["Could not parse structured risks."]
            data_sources_used = ["Database"]

    except Exception as e:
        report_content = f"Error generating report: {str(e)}"
        key_insights = []
        risk_flags = []
        data_sources_used = []

    return {
        "success": True,
        "data": {
            "supplier_id": supplier.supplier_id,
            "summary_text": report_content,
            "key_insights": key_insights,
            "risk_flags": risk_flags,
            "data_sources_used": data_sources_used,
            "generated_date": datetime.utcnow().isoformat()
        }
    }


def extract_tabular_data_locally(contents: bytes, ext: str) -> Optional[dict]:
    import io
    import csv
    
    rows = []
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            excel_file = io.BytesIO(contents)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            for r in sheet.iter_rows(values_only=True):
                rows.append([val for val in r])
        except Exception as e:
            print(f"Local Excel reading failed: {e}")
            return None
    elif ext == ".csv":
        try:
            text_content = contents.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(text_content))
            for r in reader:
                rows.append(r)
        except Exception as e:
            print(f"Local CSV reading failed: {e}")
            return None
    else:
        return None

    if not rows:
        return None

    header_mappings = {
        "product_type": ['product type', 'product_type', 'type', 'category', 'prod type', 'prod_type'],
        "sku": ['sku', 'product id', 'product_id', 'code', 'id', 'item number', 'item_number', 'item no', 'item_no', 'skucode'],
        "price": ['price', 'unit price', 'unit_price', 'sell price', 'rate', 'price/unit', 'price per unit'],
        "availability": ['availability', 'avail', 'in stock %', 'available', 'availability%', 'avail%', 'in_stock_percent'],
        "number_sold": ['number sold', 'number_sold', 'sold', 'quantity sold', 'sales', 'units sold', 'qty sold', 'sold qty', 'sold_qty', 'qty_sold'],
        "revenue": ['revenue', 'total revenue', 'turnover', 'sales amount', 'total_revenue', 'sales_amount', 'rev'],
        "customer_demographics": ['customer demographics', 'customer_demographics', 'demographics', 'gender', 'audience', 'demo'],
        "stock_level": ['stock level', 'stock_level', 'stock', 'inventory', 'quantity in stock', 'stock level qty', 'stock_qty'],
        "lead_time": ['lead time', 'lead_time', 'replenishment time', 'lead time days', 'lead_days', 'lead_time_days'],
        "order_quantity": ['order quantity', 'order_quantity', 'avg order quantity', 'order qty', 'order_qty', 'avg_order_qty'],
        "shipping_time": ['shipping time', 'shipping_time', 'transit time', 'delivery days', 'shipping_days', 'shipping_time_days'],
        "shipping_cost": ['shipping cost', 'shipping_cost', 'shipping price', 'freight cost', 'shipping_fee', 'ship_cost'],
        "shipping_carrier": ['shipping carrier', 'shipping_carrier', 'carrier', 'shipper', 'shipping_company'],
        "production_volume": ['production volume', 'production_volume', 'prod volume', 'volume', 'prod_vol', 'production_vol'],
        "manufacturing_lead_time": ['manufacturing lead time', 'manufacturing_lead_time', 'mfg lead time', 'production time', 'mfg_lead_time', 'manufacturing_time'],
        "manufacturing_cost": ['manufacturing cost', 'manufacturing_cost', 'mfg cost', 'unit cost', 'production cost', 'mfg_cost', 'cost_to_make'],
        "defect_rate": ['defect rate', 'defect_rate', 'defects', 'defect %', 'defect_rate_percentage', 'defect%', 'defects%'],
        "transportation_mode": ['transportation mode', 'transportation_mode', 'transport', 'shipping mode', 'mode', 'transport_mode'],
        "route": ['route', 'shipping route', 'transit route', 'shipping_route'],
        "inspection_result": ['inspection result', 'inspection_result', 'quality inspection', 'inspection', 'inspection_status', 'status']
    }

    header_idx = -1
    col_map = {}

    for i, r in enumerate(rows):
        if not r:
            continue
        matches = 0
        temp_map = {}
        for col_i, cell in enumerate(r):
            if cell is None:
                continue
            cell_str = str(cell).strip().lower()
            for schema_key, syns in header_mappings.items():
                if cell_str in syns:
                    temp_map[schema_key] = col_i
                    matches += 1
                    break
        if matches >= 2:
            header_idx = i
            col_map = temp_map
            break

    if header_idx == -1:
        return None

    products = []
    supplier_name = "Unknown Supplier"
    supplier_location = "Unknown"

    for i in range(header_idx):
        r = rows[i]
        if not r:
            continue
        for col_i, cell in enumerate(r):
            if cell is None:
                continue
            cell_str = str(cell).strip().lower()
            if "supplier" in cell_str or "company" in cell_str or "vendor" in cell_str:
                if col_i + 1 < len(r) and r[col_i + 1]:
                    supplier_name = str(r[col_i + 1]).strip()
                elif ":" in cell_str:
                    parts = cell_str.split(":", 1)
                    if len(parts) > 1 and parts[1].strip():
                         supplier_name = parts[1].strip().title()
            if "location" in cell_str or "address" in cell_str or "city" in cell_str:
                if col_i + 1 < len(r) and r[col_i + 1]:
                    supplier_location = str(r[col_i + 1]).strip()
                elif ":" in cell_str:
                    parts = cell_str.split(":", 1)
                    if len(parts) > 1 and parts[1].strip():
                        supplier_location = parts[1].strip().title()

    for i in range(header_idx + 1, len(rows)):
        r = rows[i]
        if not r or all(cell is None or str(cell).strip() == "" for cell in r):
            continue
        
        p_data = {}
        defaults = {
            "product_type": "skincare",
            "sku": f"SKU-{i:03d}",
            "price": 50.0,
            "availability": 80.0,
            "number_sold": 500,
            "revenue": 25000.0,
            "customer_demographics": "Female",
            "stock_level": 60.0,
            "lead_time": 15.0,
            "order_quantity": 500,
            "shipping_time": 5.0,
            "shipping_cost": 6.0,
            "shipping_carrier": "Carrier A",
            "production_volume": 800,
            "manufacturing_lead_time": 12.0,
            "manufacturing_cost": 30.0,
            "defect_rate": 2.0,
            "transportation_mode": "Road",
            "route": "Route A",
            "inspection_result": "Pass"
        }

        for key, default_val in defaults.items():
            if key in col_map:
                col_i = col_map[key]
                if col_i < len(r) and r[col_i] is not None:
                    val = r[col_i]
                    try:
                        if isinstance(default_val, int):
                            p_data[key] = int(float(str(val).replace("$", "").replace(",", "").strip()))
                        elif isinstance(default_val, float):
                            p_data[key] = float(str(val).replace("$", "").replace(",", "").replace("%", "").strip())
                        else:
                            p_data[key] = str(val).strip()
                    except Exception:
                        p_data[key] = default_val
                else:
                    p_data[key] = default_val
            else:
                p_data[key] = default_val
        products.append(p_data)

    return {
        "name": supplier_name,
        "location": supplier_location,
        "products": products
    }


async def retrieve_relevant_chunks(
    text: str,
    queries: list[str],
    api_key: Optional[str] = None,
    top_k: int = 5
) -> list[str]:
    chunk_size = 1000
    overlap = 200
    chunks = []
    start = 0
    text_stripped = text.strip()
    while start < len(text_stripped):
        end = min(start + chunk_size, len(text_stripped))
        chunk = text_stripped[start:end]
        if chunk.strip():
            chunks.append(chunk)
        start += chunk_size - overlap
        
    if not chunks:
        return []
        
    if len(chunks) <= top_k:
        return chunks

    try:
        from rag.store import embed_texts
        chunk_embeddings = embed_texts(chunks, api_key=api_key)
        
        retrieved_chunks = set()
        for query in queries:
            query_emb = embed_texts([query], api_key=api_key)[0]
            scores = []
            for i, chunk_emb in enumerate(chunk_embeddings):
                dot = sum(x * y for x, y in zip(query_emb, chunk_emb))
                norm_q = sum(x * x for x in query_emb) ** 0.5
                norm_c = sum(x * x for x in chunk_emb) ** 0.5
                sim = dot / (norm_q * norm_c) if (norm_q and norm_c) else 0.0
                scores.append((sim, chunks[i]))
            
            scores.sort(key=lambda x: x[0], reverse=True)
            for _, c in scores[:top_k]:
                retrieved_chunks.add(c)
        return list(retrieved_chunks)
        
    except Exception as e:
        print(f"Embedding-based RAG failed or rate limited ({e}). Falling back to local keyword-based RAG...")
        retrieved_chunks = set()
        for query in queries:
            terms = [t.lower() for t in query.replace(",", " ").replace(":", " ").split() if len(t) > 2]
            scores = []
            for chunk in chunks:
                chunk_lower = chunk.lower()
                score = sum(chunk_lower.count(term) for term in terms)
                scores.append((score, chunk))
            scores.sort(key=lambda x: x[0], reverse=True)
            for _, c in scores[:top_k]:
                retrieved_chunks.add(c)
        return list(retrieved_chunks)


# ── POST: Extract Supplier data from an uploaded file ──────────────
@router.post("/extract-document", response_model=dict)
async def extract_supplier_document(
    file: UploadFile = File(...),
    x_user_api_key: Optional[str] = Header(None, alias="X-User-API-Key"),
    x_user_model: Optional[str] = Header(None, alias="X-User-Model")
):
    """
    Accept an uploaded document (PDF, CSV, Excel), extract supplier name,
    location, and product rows using LLM, and return structured data.
    """
    allowed_extensions = {".pdf", ".csv", ".xlsx", ".xls"}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file format '{ext}'. Supported formats: PDF, CSV, Excel (.xlsx, .xls)"
        )

    contents = await file.read()
    
    # 1. Extract raw text based on format
    full_text = ""
    try:
        if ext == ".pdf":
            import io
            import pypdf
            pdf_file = io.BytesIO(contents)
            reader = pypdf.PdfReader(pdf_file)
            full_text = "\n".join(page.extract_text() or "" for page in reader.pages)
            
        elif ext in (".xlsx", ".xls"):
            import io
            import openpyxl
            excel_file = io.BytesIO(contents)
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            rows_text = []
            for row in sheet.iter_rows(values_only=True):
                if any(val is not None and str(val).strip() != "" for val in row):  # skip fully empty rows
                    rows_text.append(", ".join(str(val) if val is not None else "" for val in row))
            full_text = "\n".join(rows_text)
            
        elif ext == ".csv":
            full_text = contents.decode("utf-8", errors="ignore")
        else:
            full_text = contents.decode("utf-8", errors="ignore")
    except Exception as e:
        print(f"Text extraction failed: {e}")
        # If text extraction completely fails, we can check if a local tabular fallback can save us
        if ext in (".csv", ".xlsx", ".xls"):
            local_res = extract_tabular_data_locally(contents, ext)
            if local_res:
                return {
                    "success": True,
                    "data": local_res,
                    "warning": "Text extraction failed, but recovered data locally using rule-based spreadsheet parser."
                }
        raise HTTPException(status_code=400, detail=f"Failed to read file contents: {str(e)}")

    if not full_text.strip():
        # Try local fallback for spreadsheets
        if ext in (".csv", ".xlsx", ".xls"):
            local_res = extract_tabular_data_locally(contents, ext)
            if local_res:
                return {
                    "success": True,
                    "data": local_res,
                    "warning": "The file contains no readable text, but recovered data locally using rule-based spreadsheet parser."
                }
        raise HTTPException(status_code=400, detail="The uploaded file contains no readable text content.")

    # Apply in-memory RAG pipeline if document is large (e.g. > 3000 chars)
    if len(full_text) > 3000:
        queries = [
            "Supplier profile: supplier name, company name, address, location, city, country",
            "Product catalog list: items, products, SKU, price, availability, revenue, cost, demographics, lead time"
        ]
        rag_chunks = await retrieve_relevant_chunks(
            full_text, 
            queries, 
            api_key=x_user_api_key, 
            top_k=4
        )
        full_text = "\n=== RETRIEVED TEXT SEGMENTS ===\n" + "\n---\n".join(rag_chunks)
    else:
        # Limit text size for smaller files
        if len(full_text) > 20000:
            full_text = full_text[:20000] + "\n[Text truncated due to size limit...]"

    # 2. Call LLM to extract structured data
    try:
        from llm import get_llm
        from langchain_core.messages import HumanMessage
        
        # Use user model if provided, default to gemini-2.5-flash
        model_name = x_user_model or "google/gemini-2.5-flash"
        llm = get_llm(temperature=0.2, api_key=x_user_api_key, model=model_name)
        
        prompt = f"""You are an expert data extraction assistant. Your job is to extract supplier and product-level data from the raw text of an uploaded document (invoice, catalog, contract, list).
    
RAW DOCUMENT TEXT:
---
{full_text}
---

Please parse the text above and extract the following:
1. The Supplier/Company Name (or guess a reasonable company name from the document headers, default to "Unknown Supplier" if not found).
2. The Supplier/Company Location (city or country, default to "Unknown" if not found).
3. A list of all products mentioned in the document. For each product, map its properties to this specific schema:
   - product_type: type of product (e.g. cosmetics, skincare, electronics, haircare, etc.)
   - sku: SKU code or product ID (e.g. SKU-123, SKU-001)
   - price: price per unit (number)
   - availability: estimated stock availability percentage (0-100, number)
   - number_sold: quantity of products sold (number)
   - revenue: total revenue generated from this product (number, if not directly available, calculate as price * number_sold)
   - customer_demographics: primary customer demographics (e.g. Male, Female, Non-binary, All)
   - stock_level: current stock quantity (number)
   - lead_time: supplier replenishment lead time in days (number)
   - order_quantity: average order quantity (number)
   - shipping_time: shipping transit time in days (number)
   - shipping_cost: cost of shipping per unit (number)
   - shipping_carrier: name of the carrier (e.g. Carrier A, DHL, Fedex)
   - production_volume: standard production volume (number)
   - manufacturing_lead_time: production lead time in days (number)
   - manufacturing_cost: cost to manufacture per unit (number)
   - defect_rate: defect rate percentage (e.g. 1.5 for 1.5%, number)
   - transportation_mode: mode of transport (e.g. Road, Rail, Air, Sea)
   - route: shipping route name (e.g. Route A, Route B)
   - inspection_result: result of quality check (Pass, Fail, or Pending)

Instructions:
- Attempt to extract or estimate values for every field based on the text. If a field is not found in the text, use reasonable, realistic default values for that type of product (e.g. default lead_time to 10-15 days, defect_rate to 1.5, availability to 90).
- Map columns intelligently. E.g. "Unit Cost" -> manufacturing_cost, "Sales" -> number_sold, "Carrier" -> shipping_carrier.
- Return a JSON object with keys:
  - name: string
  - location: string
  - products: list of objects matching the schema above.
  
Output strictly valid JSON, no markdown blocks, no formatting wrapper, just the JSON string itself. If you output markdown blocks like ```json, the code will fail. Just output the JSON.
"""

        message = HumanMessage(content=prompt)
        response = await llm.ainvoke([message])
        content = response.content.strip()
        
        # Strip markdown if present
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()
            
        parsed_data = json.loads(content)
        
        return {
            "success": True,
            "data": {
                "name": parsed_data.get("name", "Unknown Supplier"),
                "location": parsed_data.get("location", "Unknown"),
                "products": parsed_data.get("products", [])
            }
        }

    except Exception as e:
        print(f"LLM extraction failed: {e}")
        # Try local fallback for spreadsheets
        if ext in (".csv", ".xlsx", ".xls"):
            local_res = extract_tabular_data_locally(contents, ext)
            if local_res:
                return {
                    "success": True,
                    "data": local_res,
                    "warning": f"AI extraction failed ({str(e)}), but successfully recovered data locally using rule-based spreadsheet parser."
                }
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process and extract supplier document: {str(e)}"
        )


# ── POST: Add a new supplier from product rows ──────────────────────
@router.post("/", response_model=dict)
def add_supplier(body: AddSupplierRequest, session: Session = Depends(get_session)):
    """Accept raw product rows, aggregate into supplier metrics, AI-evaluate, and save."""
    products = body.products
    if not products:
        raise HTTPException(status_code=400, detail="At least one product row is required")

    # ── 1. Aggregate product rows into supplier-level metrics ────────
    n = len(products)
    unique_skus = set()
    product_types = set()
    shipping_carriers = set()
    transportation_modes = set()
    routes = set()
    customer_demographics = set()
    inspections_pass = 0
    inspections_total = 0

    sum_price = sum_availability = sum_stock = sum_lead = 0.0
    sum_ship_time = sum_ship_cost = sum_mfg_lead = sum_mfg_cost = sum_defect = 0.0
    total_sold = total_revenue = total_order_qty = total_prod_vol = 0

    for p in products:
        unique_skus.add(p.sku)
        product_types.add(p.product_type)
        shipping_carriers.add(p.shipping_carrier)
        transportation_modes.add(p.transportation_mode)
        routes.add(p.route)
        customer_demographics.add(p.customer_demographics)

        sum_price += p.price
        sum_availability += p.availability
        sum_stock += p.stock_level
        sum_lead += p.lead_time
        sum_ship_time += p.shipping_time
        sum_ship_cost += p.shipping_cost
        sum_mfg_lead += p.manufacturing_lead_time
        sum_mfg_cost += p.manufacturing_cost
        sum_defect += p.defect_rate

        total_sold += p.number_sold
        total_revenue += p.revenue
        total_order_qty += p.order_quantity
        total_prod_vol += p.production_volume

        if p.inspection_result in ("Pass", "Fail"):
            inspections_total += 1
            if p.inspection_result == "Pass":
                inspections_pass += 1

    # Generate unique supplier id
    existing_count = len(session.exec(select(Supplier)).all())
    supplier_id = f"SUP-{existing_count + 1:03d}"

    avg_total_cost = round((sum_mfg_cost / n) + (sum_ship_cost / n), 2)
    inspection_pass_rate = round((inspections_pass / inspections_total * 100) if inspections_total > 0 else 50.0, 1)

    supplier = Supplier(
        supplier_id=supplier_id,
        name=body.name,
        location=body.location,
        product_types=json.dumps(sorted(product_types)),
        avg_price=round(sum_price / n, 2),
        avg_availability=round(sum_availability / n, 1),
        total_products_sold=total_sold,
        total_revenue=round(total_revenue, 2),
        avg_stock_level=round(sum_stock / n, 1),
        avg_lead_time=round(sum_lead / n, 1),
        total_order_quantity=total_order_qty,
        avg_shipping_time=round(sum_ship_time / n, 1),
        shipping_carriers=json.dumps(sorted(shipping_carriers)),
        avg_shipping_cost=round(sum_ship_cost / n, 2),
        total_production_volume=total_prod_vol,
        avg_manufacturing_lead_time=round(sum_mfg_lead / n, 1),
        avg_manufacturing_cost=round(sum_mfg_cost / n, 2),
        defect_rate=round(sum_defect / n, 2),
        inspection_pass_rate=inspection_pass_rate,
        transportation_modes=json.dumps(sorted(transportation_modes)),
        routes=json.dumps(sorted(routes)),
        avg_total_cost=avg_total_cost,
        customer_demographics=json.dumps(sorted(customer_demographics)),
        num_skus=len(unique_skus),
    )

    # ── 2. AI-evaluate: overall_score, risk_level, otd_percentage ────
    try:
        from llm import get_llm
        from langchain_core.messages import HumanMessage

        llm = get_llm(temperature=0.3)

        supplier_data = {
            "name": supplier.name,
            "defect_rate": supplier.defect_rate,
            "inspection_pass_rate": supplier.inspection_pass_rate,
            "avg_lead_time": supplier.avg_lead_time,
            "avg_shipping_time": supplier.avg_shipping_time,
            "avg_manufacturing_lead_time": supplier.avg_manufacturing_lead_time,
            "avg_shipping_cost": supplier.avg_shipping_cost,
            "avg_manufacturing_cost": supplier.avg_manufacturing_cost,
            "avg_total_cost": supplier.avg_total_cost,
            "total_revenue": supplier.total_revenue,
            "total_products_sold": supplier.total_products_sold,
            "avg_stock_level": supplier.avg_stock_level,
            "avg_availability": supplier.avg_availability,
        }

        prompt = f"""You are a supply chain analyst. Evaluate this supplier and compute three metrics:

SUPPLIER DATA:
{json.dumps(supplier_data, indent=2)}

Compute:
1. overall_score (0-100): Composite score weighing defect rate (lower=better), inspection pass rate (higher=better), lead times (lower=better), costs (lower=better relative to revenue), availability.
2. risk_level: "Low", "Medium", "High", or "Critical"
3. otd_percentage: Estimated on-time delivery % (0-100) based on lead/shipping/manufacturing times vs industry standards.

Return a JSON object with: "overall_score" (number), "risk_level" (string), "otd_percentage" (number).
Output ONLY valid JSON, no markdown."""

        response = llm.invoke([HumanMessage(content=prompt)])
        content = response.content
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0]
        elif "```" in content:
            content = content.split("```")[1].split("```")[0]

        ev = json.loads(content.strip())
        supplier.overall_score = ev.get("overall_score")
        supplier.risk_level = ev.get("risk_level")
        supplier.otd_percentage = ev.get("otd_percentage")
        print(f"  AI evaluated {supplier.name}: score={supplier.overall_score}, risk={supplier.risk_level}, otd={supplier.otd_percentage}%")

    except Exception as e:
        print(f"AI evaluation failed for new supplier: {e}")
        # Rule-based fallback (same logic as seed.py)
        score = 100.0
        score -= supplier.defect_rate * 10
        score -= max(0, supplier.avg_lead_time - 15) * 2
        score += supplier.inspection_pass_rate * 0.2
        score = max(0, min(100, score))
        supplier.overall_score = round(score, 1)

        if score >= 75:
            supplier.risk_level = "Low"
        elif score >= 55:
            supplier.risk_level = "Medium"
        elif score >= 35:
            supplier.risk_level = "High"
        else:
            supplier.risk_level = "Critical"
        otd = 100 - (max(0, supplier.avg_lead_time - 12) * 1.5) - (max(0, supplier.avg_shipping_time - 4) * 2)
        supplier.otd_percentage = round(max(50, min(99, otd)), 1)
        print(f"  Rule-based fallback {supplier.name}: score={supplier.overall_score}, risk={supplier.risk_level}, otd={supplier.otd_percentage}%")

    # ── 3. Save to database ──────────────────────────────────────────
    session.add(supplier)
    session.commit()
    session.refresh(supplier)

    return {"success": True, "data": supplier}


class SendMessageRequest(BaseModel):
    sender: str = "Admin"
    sender_email: str
    recipient_email: str
    recipient_phone: Optional[str] = None
    subject: str
    message: str
    sent_via: str  # "Portal", "Email", or "SMS"


class GenerateDraftRequest(BaseModel):
    subject: str


@router.post("/{supplier_id}/draft-email", response_model=dict)
async def generate_draft_email(
    supplier_id: str,
    request: GenerateDraftRequest,
    session: Session = Depends(get_session),
    x_user_api_key: Optional[str] = Header(None, alias="X-User-API-Key"),
    x_user_model: Optional[str] = Header(None, alias="X-User-Model")
):
    supplier = session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    try:
        from llm import get_llm
        from langchain_core.messages import HumanMessage

        api_key = x_user_api_key or os.getenv("OPENROUTER_API_KEY")
        model = x_user_model or "google/gemini-2.5-flash"

        llm = get_llm(api_key=api_key, model=model, temperature=0.5)

        prompt = f"""You are a supply chain analyst and communications expert at VendorVerse.
Write a professional, clear, and direct message/email to the supplier '{supplier.name}' regarding the subject: '{request.subject}'.

Use the following supplier performance data to ground the details and context if relevant (do not invent data):
- Supplier ID: {supplier.supplier_id}
- Overall Score: {supplier.overall_score if supplier.overall_score else 'N/A'}/100
- On-Time Delivery Rate: {supplier.otd_percentage if supplier.otd_percentage else 'N/A'}%
- Defect Rate: {supplier.defect_rate}%
- Inspection Pass Rate: {supplier.inspection_pass_rate}%
- Lead Time: {supplier.avg_lead_time} days
- Region: {supplier.region or 'N/A'}

Guidelines:
1. Maintain a professional, collaborative business tone.
2. Be concise and structured.
3. Do not include placeholders like '[Your Name]', '[Date]', or '[Supplier Contact]'. Keep them empty or sign off as 'VendorVerse Procurement Team'.
4. Output ONLY the email/message body content, with no markdown headers (like 'Subject:') or surrounding explanation blocks. Start directly with the greeting.
"""

        response = await llm.ainvoke([HumanMessage(content=prompt)])
        draft_content = response.content.strip()

        if draft_content.startswith("```"):
            lines = draft_content.split("\n")
            if len(lines) > 2:
                draft_content = "\n".join(lines[1:-1])

        return {"success": True, "draft": draft_content}

    except Exception as e:
        print(f"Failed to generate AI draft: {e}")
        fallback_draft = f"Dear {supplier.name} Team,\n\nWe are writing to discuss: '{request.subject}'.\n\nBased on our records, your overall performance score is {supplier.overall_score or 'N/A'}/100 with an on-time delivery rate of {supplier.otd_percentage or 'N/A'}% and a defect rate of {supplier.defect_rate}%.\n\nPlease let us know your availability to discuss this topic further.\n\nBest regards,\nVendorVerse Procurement Team"
        return {"success": True, "draft": fallback_draft}


@router.post("/{supplier_id}/messages", response_model=dict)
def send_supplier_message(
    supplier_id: str,
    request: SendMessageRequest,
    session: Session = Depends(get_session)
):
    supplier = session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    db_msg = SupplierMessage(
        supplier_id=supplier_id,
        sender=request.sender,
        sender_email=request.sender_email,
        recipient_email=request.recipient_email,
        subject=request.subject,
        message=request.message,
        sent_via=request.sent_via,
        created_at=datetime.utcnow()
    )
    session.add(db_msg)
    session.commit()
    session.refresh(db_msg)

    recipient = request.recipient_phone if request.sent_via == "SMS" else request.recipient_email
    print(f"[{request.sent_via.upper()}] Sent message from {request.sender_email} to {recipient} re: '{request.subject}'")

    return {"success": True, "data": db_msg}


@router.get("/{supplier_id}/messages", response_model=dict)
def get_supplier_messages(
    supplier_id: str,
    session: Session = Depends(get_session)
):
    supplier = session.get(Supplier, supplier_id)
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    query = select(SupplierMessage).where(SupplierMessage.supplier_id == supplier_id).order_by(SupplierMessage.created_at.desc())
    messages = session.exec(query).all()

    return {"success": True, "data": messages}
